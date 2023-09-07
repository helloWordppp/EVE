# -*- coding:utf-8 -*-

import openpyxl

class Excel:
    def __init__(self,filename,sheetname):
        '''

        :type filename: str -- excel的文件目录
        :param sheetname: str -- 选择操作的excel页名
        '''
        self.filename = filename
        try:
            self.wb = openpyxl.load_workbook(self.filename)     #生成文件对象，表示要操作的是哪个文件
        except:
            self.wb = openpyxl.Workbook()                      # 没有该文件则新建
            self.wb.save(filename)
        try:
            self.sheet = self.wb[sheetname]                #生成页对象，表示要操作的是哪个页
        except:
            self.wb.create_sheet(title=sheetname)          # 没有该sheet则新建
            self.sheet=self.wb[sheetname]
            self.wb.save(filename)

    #获取excel文件中的总行数
    def maxRow(self):
        maxrow = self.sheet.max_row
        return maxrow

    #获取excel文件中最大的列数
    def maxCol(self):
        maxcol=self.sheet.max_column
        return maxcol

    #获取某个单元格内容
    def read(self,row,col):
        info=self.sheet.cell(row=row, column=col).value
        try:
            info=str(info)
        except:
            pass
        return info

    #获取某一行数据
    def readLine(self,row):
        maxcol=self.maxCol()
        info=[]
        for i in range(1,maxcol+1):
            content=self.sheet.cell(row=row,column=i).value
            try:
                content=str(content)
            except:
                pass
            info.append(content)
        return info

    #获取所有数据
    def readall(self):
        maxrow=self.maxRow()
        info=[]
        for i in range(1,maxrow+1):
            content=self.readLine(i)
            info.append(content)
        return info

    def readCol(self,column):
        maxrow = self.maxRow()
        info=[]
        for i in range(1, maxrow + 1):
            content=self.sheet.cell(row=i, column=column).value
            try:
                content = str(content)
            except:
                pass
            info.append(content)
        return info

    # 在指定单元格写数据
    def wt(self, rownum, colnum, data):
        self.sheet.cell(row=rownum, column=colnum).value = data
        self.wb.save(self.filename)


    #在指定行写一行内容，data为列表
    def wtLine(self,row,data):
        num=len(data)
        for i in range(0,num):
            self.sheet.cell(row=row, column=i+1).value=data[i]
        self.wb.save(self.filename)


    # 写入字典数据，key为列名，value为值
    def writDict(self, infoDict):
        maxRow = self.maxRow()
        self.wtLine(maxRow+2, list(infoDict.keys()))
        self.wtLine(maxRow+3, list(infoDict.values()))


    # 写入ERE录制统计百分比数据
    def writEreRecordCount(self, title, infoDict):
        '''

        :param title: 统计数据的标题
        :param infoDict: 统计数据的字典，Key为关键字，value为数量和百分比的列表组合
        :return:
        '''
        maxRow = self.maxRow()
        # 写入标题
        self.wt(maxRow+3, 1, "title")
        self.wt(maxRow+4, 1, title)

        col = 2
        # 写入数据
        for i in infoDict.keys():
            self.wt(maxRow+3, col, i)
            self.wt(maxRow+4, col, infoDict[i][0])
            self.wt(maxRow+5, col, infoDict[i][1])
            col = col + 1








